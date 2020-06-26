def sql_format(sql, replaceDict):
    for item in replaceDict.items():
        sql = sql.replace(str(item[0]), str(item[1]))
    return sql


def getBookConnect():
    return pymysql.connect(host="xxx.com", user="xxx", password="xxx", database="xxx", charset='utf8')


import openpyxl


def read_isbn(file, colnames, sheet="Sheet1", line_num=10):
    wb = openpyxl.load_workbook(file)
    ws = wb[sheet]
    for colname in colnames:
        printISBN(colname, ws, line_num)


def printISBN(colname, ws, line_num=10):
    print(f"\n\n-- Current Colnum:{colname}")
    isbn_list = ws[colname]
    count = 1
    length = len(isbn_list)
    for isbn in isbn_list:
        isbn = isbn.value
        isbn = str(isbn).replace("-", "")

        if count == length:
            print(f"{isbn}", )
        else:
            print(f"{isbn},", end="")

        if count % line_num == 0 and count != length:
            print("")
        count = count + 1


# ----------------------------------------------

sql_temp_update = '''
UPDATE book_keyword bk
SET is_delete = 1
WHERE bk.book_id = {bookID} AND bk.is_delete = 0 AND classify_id <> 0;\n
'''

sql_temp_insert = '''
INSERT INTO `book_keyword` (`keyword_id`, `book_group_id`, `book_id`, `channel_id`,           `classify_id`, `rank`, `set_type`, `is_delete`, `create_user`, `create_time`, `update_user`, `update_time`, `is_warehouse`, `warehouse_id`, `is_edit`)(
SELECT                      `keyword_id`, `book_group_id`, `book_id`, `channel_id`,  {classifyID} `classify_id`, `rank`, `set_type`, `is_delete`, `create_user`, `create_time`, `update_user`, `update_time`, `is_warehouse`, `warehouse_id`, `is_edit`
FROM
    book_keyword bk
WHERE
    bk.book_id = {bookID} AND bk.is_delete = 0 AND classify_id = 0
);\n
'''

bookIds = [2966350, 2966351, 2984439, 4120189, 4534243, 4535159, 4550713, 4551599, 4551655, 4782047,
           4782146, 4782235, 4782686, 4786228, 4786768, 4792270, 4792271, 4910500, 4910828, 4910844,
           4911121, 4911552, 4911584, 4911664, 4911704, 4911709, 4911715, 4912281, 4912282, 4912755,
           4912833, 4912834, 4912839, 4912840, 4912842, 4912843, 4912844, 4912846, 4912848, 4912859,
           4912860, 4912861, 4912862, 4912863, 4912864, 4912865, 4912866, 4912867, 4912869, 4912870,
           4912931, 4912958, 4912971, 4912972, 4912973, 4913072, 4913843, 4913903, 4915161, 4956845,
           4994546, 4994830, 4994905, 5003814, 5011313, 5011427, 5011432]

sql_get_classify_id = '''
SELECT DISTINCT classify_id FROM book_keyword bk
WHERE bk.book_id = {bookID}  and classify_id <> 0;
'''


def get_classify_ids(conn, bookId):
    sql = sql_format(sql_get_classify_id, {"{bookID}": bookId})
    cursor = conn.cursor()
    cursor.execute(sql)
    data = cursor.fetchall()
    cursor.close()
    return [x[0] for x in data]


with open("update.sql", "w") as f:
    for bookId in bookIds:
        f.write(sql_format(sql_temp_update, {"{bookID}": bookId}))

with open("insert.sql", "w") as f, getBookConnect() as conn:
    for bookId in bookIds:
        classify_ids = get_classify_ids(conn, bookId)
        for classify_id in classify_ids:
            f.write(sql_format(sql_temp_insert, {"{bookID}": bookId, "{classifyID}": classify_id}))

# ----------------------------------------------

import math
import pymysql
import datetime
from openpyxl import Workbook


def calc_entroy(name: str):
    chars = {}
    for c in name:
        if c != '群':
            if c not in chars:
                chars[c] = 1
            else:
                chars[c] = chars[c] + 1
    sumValue = 0
    for key in chars:
        sumValue += chars[key]
    value = 0
    for key in chars:
        # value = value + chars[key]* math.log(chars[key]/sumValue)
        value = value + chars[key] * math.log(chars[key])
    value = value / sumValue
    return value


stop_word = ["的", "地", "我"]


def calc_base(name: str):
    # 1. 包含三个数字, 字母
    # 2. 包含 社群
    # 3. 包含任意相同的三个字符 例如  心心心心心心心心心心1群
    # 4. 包含停止词和其他无意义词汇   的 地 我
    chars = {}
    c_alpha_and_num = 0

    if "社群" in name:
        return "group"

    for c in name:
        if 'a' < c < 'z' or 'A' < c < 'Z' or '0' < c < '9':
            c_alpha_and_num = c_alpha_and_num + 1
        if c != '群':
            if c not in chars:
                chars[c] = 1
            else:
                chars[c] = chars[c] + 1
        if c in stop_word:
            return "stop"
    if c_alpha_and_num >= 4:
        return "alpha"
    for key in chars:
        if chars[key] >= 3:
            return "multi"

    return None


columns = ['id', 'classify_id', 'group_name', 'weixin_qrcode_id',
           'weixin_group_id', 'user_number', 'create_time']
sql = f"SELECT {', '.join(columns)} FROM book_group_qrcode WHERE user_number < 16 AND create_time < '2019-11-23';"


def insertHead(sheet):
    for i, name in enumerate(columns):
        sheet.cell(row=1, column=i + 1, value=name)
    sheet.cell(row=1, column=i + 2, value='熵值')
    sheet.cell(row=1, column=i + 3, value='属性')


wb = Workbook()
sheet = wb.create_sheet("群名称熵值分析", index=0)

with getBookConnect() as cursor:
    cursor.execute(sql)
    resultSet = cursor.fetchall()
    insertHead(sheet)
    for i, item in enumerate(resultSet, start=2):
        for col, name in enumerate(columns):
            sheet.cell(row=i, column=col + 1, value=item[col])
        sheet.cell(row=i, column=col + 2, value=calc_entroy(item[2]))
        sheet.cell(row=i, column=col + 3, value=calc_base(item[2]))

wb.save(f"群名称熵值分析({datetime.datetime.now().strftime('%Y-%m-%d-%H')}).xlsx")
